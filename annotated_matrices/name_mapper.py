import re
from pathlib import Path
from copy import copy

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


def parse_sample_id(sample_id):
    "Funkce parse_sample_id zpracovává hodnoty Sample ID ze souboru se jmény, "
    "ze vstupního řetězce extrahuje textový prefix a číselnou část identifikátoru a vrací je ve standardizované podobě."
    if pd.isna(sample_id):
        return None

    s = str(sample_id).strip()
    m = re.search(r'([A-Za-z]+)[^0-9]*([0-9]+)', s)
    if not m:
        return None

    prefix = m.group(1).upper()
    num = int(m.group(2))
    return prefix, num


def parse_matrix_id(matrix_id):
    "Funkce parse_matrix_id zpracovává identifikátory vzorků z genotypové matice,"
    " extrahuje z nich textový prefix a číselnou část ID a umožňuje jejich propojení se záznamy ve vstupním souboru se jmény"
    if matrix_id is None:
        return None

    s = str(matrix_id).strip()
    parts = [p.strip() for p in s.split("-") if p.strip()]

    last_digit_idx = None
    num = None

    for i in range(len(parts) - 1, -1, -1):
        token = parts[i]
        if token.isdigit():
            num = int(token)
            last_digit_idx = i
            break

    if num is None:
        return None

    prefix = None
    for j in range(last_digit_idx - 1, -1, -1):
        token = parts[j]
        if token.isalpha():
            prefix = token.upper()
            break

    if prefix is None:
        return None

    return prefix, num


def insert_genotype_names(matrix_path, names_path):
    """Funkce key_names načte Excelovou matici a soubor se jmény (names_path), 
    vytvoří mapování mezi Sample ID a dvojicí Genotype Name + Description a poté tuto informaci doplní do každého listu matice. 
    Zároveň ve všech listech vloží nové sloupce Genotype Name a Description, zachová strukturu a formátování tabulek a výsledný soubor uloží.
    """

    # --- načtení names ---
    names_df = pd.read_excel(names_path)

    mapping = {}
    for _, row in names_df.iterrows():
        parsed = parse_sample_id(row.get("Sample ID"))
        if not parsed:
            continue

        key = parsed
        genotype = row.get("Genotype Name")
        description = row.get("Description")
        mapping[key] = (genotype, description)

    # --- načtení matrix ---
    wb = load_workbook(matrix_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        max_row = ws.max_row
        max_col = ws.max_column

        # 1) Záloha a odmergování všech merged cells
        original_merged = list(ws.merged_cells.ranges)
        for mr in original_merged:
            ws.unmerge_cells(str(mr))

        # 2) POSUN VŠEHO OD SLOUPCE B O 2 SLOUPCE DOPRAVA
        old_max_col = max_col
        for r in range(1, max_row + 1):
            for c in range(old_max_col, 1, -1):   # od max_col do 2 (A necháváme)
                src = ws.cell(row=r, column=c)
                dst = ws.cell(row=r, column=c + 2)

                # hodnota
                dst.value = src.value

                # formátování
                dst.font = copy(src.font)
                dst.border = copy(src.border)
                dst.fill = copy(src.fill)
                dst.number_format = src.number_format
                dst.alignment = copy(src.alignment)

                # původní buňku vyčistíme
                src.value = None

        # 3) Znovu vytvořit merged cells, ale ty od sloupce B dál posunout o 2
        from openpyxl.utils import get_column_letter, column_index_from_string

        for mr in original_merged:
            min_row, max_row_m = mr.min_row, mr.max_row
            min_col, max_col_m = mr.min_col, mr.max_col

            if min_col >= 2:
                new_min_col = min_col + 2
                new_max_col = max_col_m + 2
            else:
                new_min_col = min_col
                new_max_col = max_col_m

            ws.merge_cells(
                start_row=min_row,
                end_row=max_row_m,
                start_column=new_min_col,
                end_column=new_max_col,
            )

        # 4) Nové sloupce B a C – border + tučný font + BÍLÉ POZADÍ
        white_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")  # <<< bílá výplň

        max_row = ws.max_row  # pro jistotu znovu
        for r in range(1, max_row + 1):
            src_cell = ws[f"A{r}"]
            for col in ("B", "C"):
                dst_cell = ws[f"{col}{r}"]
                if src_cell.border:
                    dst_cell.border = copy(src_cell.border)
                dst_cell.font = Font(bold=True)
                dst_cell.fill = white_fill  # <<< tady se explicitně smaže původní barva

        # 5) Hlavička nových sloupců (řádek 7)
        header_row = 7
        ws[f"B{header_row}"] = "Genotype Name"
        ws[f"C{header_row}"] = "Description"

        # 6) Vyplnit Genotype / Description podle Sample ID ve sloupci A (od řádku 8)
        start_data_row = 8
        for r in range(start_data_row, max_row + 1):
            cell_value = ws[f"A{r}"].value
            parsed = parse_matrix_id(cell_value)
            if parsed and parsed in mapping:
                genotype, description = mapping[parsed]
                ws[f"B{r}"] = genotype
                ws[f"C{r}"] = description

    OUT_FILE = "All_Traits_matrix_annotated.xlsx"
    out_folder = Path("annotated_matrices")
    out_folder.mkdir(parents=True, exist_ok=True)
    out_path = out_folder / OUT_FILE

    wb.save(out_path)
    # ?return out_path
    print(f"Hotovo! Vytvořen soubor: {out_path}")
