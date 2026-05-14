from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from copy import copy as _copy
import re

from gene_rules import GENE_RULES  # {"rules": {...}, "exceptions": {...}}
from special_status import eval_special_h3_h2, eval_special_single


def evaluate_gene_status(vstupni_soubor: str):
    """
    Pro každý list:
      - najde sekce podle sloučených buněk v 1. řádku (kromě 'Gene_name'),
      - za každou sekci vloží 1 prázdný sloupec,
      - do 7. řádku zapíše 'Status' do POSLEDNÍHO sloupce sekce (o 1 vlevo od mezery),
      - od řádku 8 vyhodnocuje hodnoty v sekci.

    REF/ALT se určují podle gene_rules:

    Standardní výsledek sekce:
      - pokud se v sekci najde ALT → 'non-functional'
      - jinak, pokud se v sekci objeví Het / N / alt2 a není ALT → 'unknown'
      - jinak → 'functional'

    Speciální výjimky (mají přednost):
      - H3_delays(P) / H2_speedsup(P) / H1_Ref  → kombinace pozic (viz special_status_rules.py)
      - highCd / lowCd → pozice 4966222: REF → highCd, ALT → lowCd

    Pokud se trefí některé speciální pravidlo, nepoužije se už
    běžné non-functional/unknown/functional.
    """

    OUT_FILE = "All_Traits_status_evaluated.xlsx"
    out_folder = Path("evaluated_status_matrices")
    out_folder.mkdir(parents=True, exist_ok=True)
    out_path = out_folder / OUT_FILE

    wb = load_workbook(vstupni_soubor, data_only=False)

    HEADER_ROW = 1
    STATUS_ROW = 7  # kam píšeme text "Status"

    RULES = GENE_RULES.get("rules", {})
    EXCEPTIONS = GENE_RULES.get("exceptions", {})

    # Styl pro nápis "Status"
    bold_black_font = Font(bold=True, color="000000")

    # Stylování pro status
    blue_fill = PatternFill(start_color="FFCCE5FF", end_color="FFCCE5FF", fill_type="solid")
    gray_fill = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")

    # Normalizace pro porovnání (ref/alt)
    def norm(v):
        if v is None:
            return ""
        s = str(v).strip()
        # odstraň uvozovky na začátku/konci, ale NEMĚŇ velikost písmen
        s = re.sub(r'^[\'"„“‚‘]|[\'"”“’]$', '', s)
        return s.strip()

    for ws in list(wb.worksheets):
        max_row = ws.max_row
        max_col = ws.max_column
        if max_col == 0:
            continue

        # --- Najdi řádek s 'variant_position' v 1. sloupci (pokud existuje)
        VARIANT_ROW = None
        for r in range(1, max_row + 1):
            v = ws.cell(row=r, column=1).value
            if v is None:
                continue
            if norm(v).lower() == "variant_position":
                VARIANT_ROW = r
                break

        # --- 1) Najdi sekce v 1. řádku, KROMĚ 'Gene_name'
        merged_on_row1 = [
            mr for mr in ws.merged_cells.ranges
            if mr.min_row <= HEADER_ROW <= mr.max_row
        ]

        covered_cols = set()
        all_sections = []

        # sloučené bloky
        for mr in merged_on_row1:
            anchor = ws.cell(row=mr.min_row, column=mr.min_col)
            header_val = norm(anchor.value)
            if header_val and header_val != "Gene_name":
                all_sections.append((mr.min_col, mr.max_col))
                covered_cols.update(range(mr.min_col, mr.max_col + 1))

        # single buňky mimo merge
        for c in range(1, max_col + 1):
            if c in covered_cols:
                continue
            val = ws.cell(row=HEADER_ROW, column=c).value
            header_val = norm(val)
            if header_val and header_val != "Gene_name":
                all_sections.append((c, c))

        if not all_sections:
            continue

        all_sections.sort(key=lambda ab: ab[0])  # Zleva doprava
        sections = all_sections

        # --- 3) Mapování: starý sloupec -> nová pozice (po vložení mezer)
        ends = [end for _, end in sections]

        def offset_for_col(c: int) -> int:
            return sum(1 for e in ends if e < c)

        col_map = {old: old + offset_for_col(old) for old in range(1, max_col + 1)}

        # kde budou PRÁZDNÉ sloupce (za každou sekci)
        spacer_cols = [end + offset_for_col(end + 1) + 1 for (_, end) in sections]
        # kde má být nápis 'Status' — poslední sloupec sekce (tj. o 1 vlevo od mezery)
        status_targets = [max(1, sc - 1) for sc in spacer_cols]

        # --- 4) Nový list + kopie dat a stylů
        insert_idx = wb.worksheets.index(ws)
        tmp_name_base = (ws.title[:20] + "_tmp")[:25]
        tmp_name = tmp_name_base
        suffix = 1
        while any(sh.title == tmp_name for sh in wb.worksheets):
            suffix += 1
            tmp_name = f"{tmp_name_base}_{suffix}"
        ws_new = wb.create_sheet(title=tmp_name, index=insert_idx)

        # šířky sloupců
        for old_col in range(1, max_col + 1):
            new_col = col_map[old_col]
            old_letter = get_column_letter(old_col)
            new_letter = get_column_letter(new_col)
            width = (
                ws.column_dimensions.get(old_letter, None).width
                if old_letter in ws.column_dimensions
                else None
            )
            if width is not None:
                ws_new.column_dimensions[new_letter].width = width

        # výšky řádků
        for r, rd in ws.row_dimensions.items():
            if rd.height is not None:
                ws_new.row_dimensions[r].height = rd.height

        # kopie buněk (hodnoty + styly)
        for r in range(1, max_row + 1):
            for old_col in range(1, max_col + 1):
                new_col = col_map[old_col]
                src = ws.cell(row=r, column=old_col)
                dst = ws_new.cell(row=r, column=new_col)
                dst.value = src.value
                dst.number_format = src.number_format
                if src.has_style:
                    dst.font = _copy(src.font)
                    dst.fill = _copy(src.fill)
                    dst.border = _copy(src.border)
                    dst.alignment = _copy(src.alignment)
                    dst.protection = _copy(src.protection)

        # zapiš 'Status' do 7. řádku na konec každé sekce
        for col in status_targets:
            cell = ws_new.cell(row=STATUS_ROW, column=col)
            cell.value = "Status"
            cell.font = bold_black_font

        # freeze panes
        ws_new.freeze_panes = ws.freeze_panes

        # --- 5) Vyhodnocení statusu (special + standardní logika)
        for row in range(STATUS_ROW + 1, max_row + 1):  # od řádku 8 dolů
            for (sec_start, sec_end), status_col in zip(sections, status_targets):

                # --- SPECIÁLNÍ VÝJIMKY ---
                # pos_results: pozice -> "REF" / "ALT"
                pos_results = {}
                single_status = None  # např. "highCd" / "lowCd"

                if VARIANT_ROW is not None:
                    for col in range(col_map[sec_start], col_map[sec_end] + 1):
                        pos_val = ws_new.cell(row=VARIANT_ROW, column=col).value
                        pos_key = norm(pos_val)
                        if not pos_key:
                            continue

                        # získej ref/alt pro tuto pozici
                        ref_name = alt_name = None
                        if pos_key in EXCEPTIONS:
                            ref_name, alt_name = EXCEPTIONS[pos_key]
                        elif pos_key in RULES:
                            ref_name, alt_name = RULES[pos_key]
                        else:
                            continue

                        ref_name = "" if ref_name is None else norm(ref_name)
                        alt_name = "" if alt_name is None else norm(alt_name)

                        cell_val = norm(ws_new.cell(row=row, column=col).value)
                        if cell_val == "":
                            continue

                        # zjisti, jestli tato buňka odpovídá REF nebo ALT
                        if cell_val == ref_name:
                            pos_results[pos_key] = "REF"
                            cand = eval_special_single(pos_key, "REF")
                            if cand is not None:
                                single_status = cand
                        elif cell_val == alt_name:
                            pos_results[pos_key] = "ALT"
                            cand = eval_special_single(pos_key, "ALT")
                            if cand is not None:
                                single_status = cand

                status_cell = ws_new.cell(row=row, column=status_col)

                # 1) Nejprve kombinace H3/H2/H1_Ref
                status_from_special = eval_special_h3_h2(pos_results)
                if status_from_special is not None:
                    status_cell.value = status_from_special

                    # barvit jen H3/H2 (ALT komba), H1_Ref necháme bez barvy
                    if status_from_special in ("H3_delays(P)", "H2_speedsup(P)"):
                        status_cell.fill = blue_fill

                    continue  # na další sekci (přeskoč standardní logiku)

                # 2) Pokud není H3/H2/H1, ale máme single_status (highCd/lowCd, apod.)
                if single_status is not None:
                    status_cell.value = single_status

                    # Barvit jen ALT (lowCd), REF (highCd) nechat bez barvy
                    if single_status == "lowCd":
                        status_cell.fill = blue_fill

                    continue  # na další sekci

                # --- STANDARDNÍ LOGIKA (non-functional / unknown / functional) ---
                found_alt = False
                found_unknown = False  # Het nebo N v sekci

                for col in range(col_map[sec_start], col_map[sec_end] + 1):
                    cell_value = ws_new.cell(row=row, column=col).value
                    cell_norm = norm(cell_value)

                    # prázdná buňka → ignorovat
                    if cell_norm == "":
                        continue

                    # Het / N / alt2 → unknown (case-insensitive jen pro tyhle tři)
                    if cell_norm.lower() in ("het", "n", "alt2"):
                        found_unknown = True
                        continue

                    # --- REF/ALT přes gene_rules ---
                    alt_name = ref_name = None
                    if VARIANT_ROW is not None:
                        pos_val = ws_new.cell(row=VARIANT_ROW, column=col).value
                        pos_key = norm(pos_val)
                        if pos_key:
                            if pos_key in EXCEPTIONS:
                                ref_name, alt_name = EXCEPTIONS[pos_key]
                            elif pos_key in RULES:
                                ref_name, alt_name = RULES[pos_key]

                    if ref_name is not None or alt_name is not None:
                        ref_name = "" if ref_name is None else norm(ref_name)
                        alt_name = "" if alt_name is None else norm(alt_name)

                        # ALT podle gene_rules
                        if alt_name and cell_norm == alt_name:
                            found_alt = True
                            break
                        # REF podle gene_rules → ignorovat (pokračovat)
                        if ref_name and cell_norm == ref_name:
                            continue
                        # jiná hodnota než REF/ALT → ignorovat
                        continue

                    # pokud pro daný sloupec neexistuje pravidlo v gene_rules,
                    # tahle buňka se prostě ignoruje

                # zapiš výsledek + barvu
                if found_alt:
                    status_cell.value = "non-functional"
                    status_cell.fill = blue_fill
                elif found_unknown:
                    status_cell.value = "unknown"
                    status_cell.fill = gray_fill
                else:
                    status_cell.value = "functional"

        # --- 6) Přemapuj merged oblasti
        for mr in ws.merged_cells.ranges:
            a_row, a_col, b_row, b_col = mr.min_row, mr.min_col, mr.max_row, mr.max_col
            new_a_col = col_map.get(a_col, a_col)
            new_b_col = col_map.get(b_col, b_col)
            left, right = sorted([new_a_col, new_b_col])
            ws_new.merge_cells(
                start_row=a_row, start_column=left, end_row=b_row, end_column=right
            )

        # --- 7) Nahraď starý list
        wb.remove(ws)
        ws_new.title = ws.title

    wb.save(out_path)
    print(f"Hotovo! Vytvořen soubor: {out_path}")
