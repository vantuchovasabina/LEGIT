from pathlib import Path
from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from gene_rules import GENE_RULES


def format_trait_matrix(soubor):
    "Funkce color_alt_in_all_sheets načte Excel soubor s maticí traitů a ve všech listech obarví buňky podle typu genotypu. "
    "Na základě pravidel z GENE_RULES zvýrazňuje alternativní alelu (ALT) modře, hodnoty Het oranžově a hodnoty N nebo ALT2 šedě. "
    "Zároveň u vybraných pozic vloží pod danou pozici řádek s popiskem Marker (posune data o řádek dolů) a u listu Flowering_time_&_maturity přeskupí sloupce/sekce podle pořadí genů. "
    
    OUT_FILE = "All_Traits_matrix_formatted.xlsx"
    out_folder = Path("formatted_matrices")
    out_folder.mkdir(parents=True, exist_ok=True)
    out_path = out_folder / OUT_FILE

    wb = load_workbook(soubor)

    def _value_with_merged(ws, row, col):
        cell = ws.cell(row=row, column=col)
        if cell.value is not None:
            return cell.value
        coord = cell.coordinate
        for mrange in ws.merged_cells.ranges:
            if coord in mrange:
                tl = ws.cell(row=mrange.min_row, column=mrange.min_col)
                return tl.value
        return None

    blue_fill   = PatternFill(start_color="FFCCE5FF", end_color="FFCCE5FF", fill_type="solid")
    orange_fill = PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid")
    gray_fill   = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")

    RULES = GENE_RULES.get("rules", {})
    EXCEPTIONS = GENE_RULES.get("exceptions", {})

    # --- 1) Barvení na všech listech ---
    for ws in wb.worksheets:
        # najdi řádky ALT / REF / VARIANT_POSITION v 1. sloupci
        alt_row = ref_row = var_row = None
        for r in range(1, ws.max_row + 1):
            v = _value_with_merged(ws, r, 1)
            if v is None:
                continue
            s = str(v).strip().lower()
            if s == "alt":
                alt_row = r
            elif s == "ref":
                ref_row = r
            elif s == "variant_position":
                var_row = r

        # -------- MODRÉ BARVENÍ (ALT) podle GENE_RULES --------
        if var_row is not None and alt_row is not None:
            for c in range(2, ws.max_column + 1):
                var_val = _value_with_merged(ws, var_row, c)
                if var_val is None:
                    continue

                pos_key = str(var_val).strip()

                # 1) priorita: EXCEPTIONS
                if pos_key in EXCEPTIONS:
                    ref_name, alt_name = EXCEPTIONS[pos_key]
                # 2) běžná pravidla
                elif pos_key in RULES:
                    ref_name, alt_name = RULES[pos_key]
                else:
                    continue  # žádné pravidlo pro tuto pozici

                ref_name = "" if ref_name is None else str(ref_name).strip()
                alt_name = "" if alt_name is None else str(alt_name).strip()
                if not (ref_name or alt_name):
                    continue

                # projdeme datové řádky POD ALT
                for r in range(alt_row + 1, ws.max_row + 1):
                    cell = ws.cell(row=r, column=c)
                    if cell.value is None:
                        continue
                    val = str(cell.value).strip()

                    # REF -> nic
                    if ref_name and val == ref_name:
                        continue

                    # ALT -> modře
                    if alt_name and val == alt_name:
                        cell.fill = blue_fill
                        continue
                    # ostatní ignorujeme

        # -------- ORANŽOVÁ / ŠEDÁ (Het / N / alt2) --------
        data_start = (alt_row + 1) if alt_row is not None else 1
        for r in range(data_start, ws.max_row + 1):
            for c in range(2, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                if cell.value is None:
                    continue
                low = str(cell.value).strip().lower()
                if low == "het":
                    cell.fill = orange_fill
                elif low in ("n", "alt2"):
                    cell.fill = gray_fill

        # --- Marker pod vybrané pozice ---
        TARGET_POSITIONS = {
            "16046818",
            "14725561", "14726160",
            "44049795", "36600513",
            "7775970", "7776045",
            "29916524", "29966815",
            "15681367", "15680659",
            "46927166", "46927184",
        }

        hits = []
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                val = _value_with_merged(ws, r, c)
                if val is None:
                    continue
                if str(val).strip() in TARGET_POSITIONS:
                    hits.append((r, c))

        for r, c in hits:
            target_r = r + 1
            if target_r <= ws.max_row and ws.cell(row=target_r, column=c).value not in (None, ""):
                for rr in range(ws.max_row, target_r - 1, -1):
                    src = ws.cell(row=rr,   column=c)
                    dst = ws.cell(row=rr+1, column=c)
                    dst.value = src.value
                    try: dst.fill = copy(src.fill)
                    except Exception: pass
                    try: dst.font = copy(src.font)
                    except Exception: pass
                    try: dst.border = copy(src.border)
                    except Exception: pass
                    try: dst.alignment = copy(src.alignment)
                    except Exception: pass
                    dst.number_format = src.number_format
                    src.value = None

            src_hdr = ws.cell(row=r, column=c)
            dst_cell = ws.cell(row=target_r, column=c)
            dst_cell.value = "Marker"
            try: dst_cell.font = copy(src_hdr.font)
            except Exception: pass
            try: dst_cell.alignment = copy(src_hdr.alignment)
            except Exception: pass
            try: dst_cell.border = copy(src_hdr.border)
            except Exception: pass
            try: dst_cell.fill = copy(src_hdr.fill)
            except Exception: pass
            dst_cell.number_format = src_hdr.number_format

    # --- 3) Přeskládání SEKCI podle 'Gene_name' ---
    target_sheet_name = "Flowering_time_&_maturity"
    if target_sheet_name in wb.sheetnames:
        ws = wb[target_sheet_name]
        header_row = None
        for r in range(1, ws.max_row + 1):
            v = _value_with_merged(ws, r, 1)
            if v is not None and str(v).strip().lower() == "gene_name":
                header_row = r
                break

        if header_row is not None:
            desired = [
                "E1", "E1LB", "E1LA", "E2", "E3", "E4", "E6/J", "E7?", "E9/FT2a",
                "FT1A", "FT1B", "FULb", "GmAP1d", "GmFRL1", "Tof12", "Tof5/FULc"
            ]
            sections = []
            cur_name = None
            cur_start = None
            for c in range(2, ws.max_column + 1):
                name = _value_with_merged(ws, header_row, c)
                name = None if name is None else str(name).strip()
                if name != cur_name:
                    if cur_name is not None:
                        sections.append((cur_name, cur_start, c - 1))
                    cur_name = name
                    cur_start = c
            if cur_name is not None:
                sections.append((cur_name, cur_start, ws.max_column))

            name_to_sections = {}
            for name, s, e in sections:
                name_to_sections.setdefault(name, []).append((name, s, e))
            ordered_sections = []
            for g in desired:
                if g in name_to_sections:
                    ordered_sections.extend(name_to_sections.pop(g))
            for lst in name_to_sections.values():
                ordered_sections.extend(lst)

            same = True
            idx = 2
            for _, s, e in ordered_sections:
                width = e - s + 1
                if s != idx:
                    same = False
                    break
                idx += width

            if not same:
                tmp = wb.create_sheet(title=f"{target_sheet_name}__tmp")
                col_map = {1: 1}
                out_c = 2
                for _, s, e in ordered_sections:
                    for c in range(s, e + 1):
                        col_map[c] = out_c
                        out_c += 1

                for r in range(1, ws.max_row + 1):
                    for old_c, new_c in col_map.items():
                        src = ws.cell(row=r, column=old_c)
                        dst = tmp.cell(row=r, column=new_c)
                        dst.value = src.value
                        try: dst.fill = copy(src.fill)
                        except Exception: pass
                        try: dst.font = copy(src.font)
                        except Exception: pass
                        try: dst.border = copy(src.border)
                        except Exception: pass
                        try: dst.alignment = copy(src.alignment)
                        except Exception: pass
                        dst.number_format = src.number_format

                for old_c, new_c in col_map.items():
                    sL = get_column_letter(old_c)
                    dL = get_column_letter(new_c)
                    src_dim = ws.column_dimensions.get(sL)
                    if src_dim is not None and src_dim.width:
                        tmp.column_dimensions[dL].width = src_dim.width

                for m in ws.merged_cells.ranges:
                    if m.min_col in col_map and m.max_col in col_map:
                        tmp.merge_cells(
                            start_row=m.min_row,
                            start_column=col_map[m.min_col],
                            end_row=m.max_row,
                            end_column=col_map[m.max_col],
                        )

                wb.remove(ws)
                tmp.title = target_sheet_name
    

    # --- 4) Seřazení listů (adjacency pravidla) ---
    def _move_sheet_before(wb, sheet_name, before_name):
        """Move sheet_name to be immediately before before_name (if both exist)."""
        if sheet_name not in wb.sheetnames or before_name not in wb.sheetnames:
            return

        sheets = wb._sheets  # list of Worksheet objects
        # najdi objekty
        s_obj = wb[sheet_name]
        b_obj = wb[before_name]

        # pokud je to už hned vedle správně, nic nedělej
        s_idx = sheets.index(s_obj)
        b_idx = sheets.index(b_obj)
        if s_idx == b_idx - 1:
            return

        # vyndej a vlož
        sheets.remove(s_obj)
        b_idx = sheets.index(b_obj)  # po remove přepočítat index
        sheets.insert(b_idx, s_obj)

    # Seed_coat_Green vedle Stay_Green
    _move_sheet_before(wb, "Stay_Green", "Seed_coat_Green")

    # Stem_termination vedle Semi-determinate_stem
    _move_sheet_before(wb, "Stem_termination", "Semi-determinate_stem")

    # Photosensitivity vedle Flowering_time_&_maturity
    _move_sheet_before(wb, "Flowering_time_&_maturity", "Photosensitivity")

    # --- Skupina Big_seed / Protein / Oil ....
    _move_sheet_before(wb, "Protein_Oil", "Oil")
    _move_sheet_before(wb, "Protein_oil_content", "Protein_Oil")
    _move_sheet_before(wb, "Big_seed_Protein_Oil", "Protein_oil_content")
    _move_sheet_before(wb, "Big_seed", "Big_seed_Protein_Oil")



    wb.save(out_path)
    print(f"Hotovo! Vytvořen soubor: {out_path}")
    return str(out_path)
