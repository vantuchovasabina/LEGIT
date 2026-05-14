from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, PatternFill
from gene_rules import GENE_RULES  # <- DŮLEŽITÉ


def phenotype_function(vstupni_soubor):
    """
    Funkce pridej_phenotype_sloupce načte Excelový soubor s genotypovou maticí, 
    na každém listu přidá nový sloupec phenotype a pro každý vzorek v řádcích vypočítá výsledný 
    fenotyp na základě hodnot REF/ALT a pravidel definovaných v GENE_RULES. Každý trait se vyhodnocuje podle vlastních pravidel.
    Výsledek zároveň barevně zvýrazní podle typu fenotypu

    1) Prázdné buňky ("") se ignorují úplně pro všechny listy KROMĚ Sucrose-like listů
       (tam se prázdno počítá do totalu pro procenta).
       => tzn. prázdná buňka neovlivní žádné flagy (all_ref, all_called, atd.) mimo sucrose-like.

    2) Globální priorita "Status":
       - Projde se celý 7. řádek a najdou se sloupce s hlavičkou "Status"
       - Pro každý datový řádek se zkontrolují všechny status buňky v daném řádku
       - Pokud je kdekoliv "unknown", nastaví se phenotype = "unknown" (šedě) a řádek se přeskočí
    """

    HEADER_ROW = 7
    FIRST_DATA_ROW = 8

    RULES = GENE_RULES.get("rules", {})
    EXCEPTIONS = GENE_RULES.get("exceptions", {})

    # ohraničení buněk
    thick = Side(border_style="medium", color="000000")
    cell_border = Border(left=thick, right=thick, top=thick, bottom=thick)

    # výplně
    light_green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    pastel_blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    pastel_pink_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    pastel_purple_fill = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")

    # Pubsecence_color barvy
    tawny_fill = PatternFill(start_color="C49A6C", end_color="C49A6C", fill_type="solid")
    light_tawny_fill = PatternFill(start_color="F8E5C1", end_color="F8E5C1", fill_type="solid")

    # Pod_color barvy
    tan_fill = PatternFill(start_color="F3E0C2", end_color="F3E0C2", fill_type="solid")
    brown_fill = PatternFill(start_color="A67C52", end_color="A67C52", fill_type="solid")
    light_black_fill = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")

    # Flowering_time_&_maturity barvy
    ultra_early_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    very_early_fill  = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    early_fill       = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    mid_early_fill   = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    late_fill        = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    # ---- MAPOVÁNÍ pro Sucrose-like listy ----
    SUCROSE_LIKE_MAP = {
        "Sucrose": {
            "mid_max": 50.0,
            "mid_inclusive": True,
            "zero": ("normal", None),
            "mid": ("slightly increased sucrose content", pastel_blue_fill),
            "high": ("Potentially lower sucrose content", pastel_pink_fill),
        },
        "Anti-nutritional": {
            "mid_max": 50.0,
            "mid_inclusive": True,
            "zero": ("normal", None),
            "mid": ("likely lower", pastel_blue_fill),
            "high": ("lower", pastel_pink_fill),
        },
        # "Oil": {
        #     "mid_max": 50.0,
        #     "mid_inclusive": True,
        #     "zero": ("normal", None),
        #     "mid": ("slightly", pastel_blue_fill),
        #     "high": ("increased", pastel_pink_fill),
        # },
        "Salt_tolerance": {
            "mid_max": 50.0,
            "mid_inclusive": True,
            "zero": ("normal", None),
            "mid": ("slightly tolerant", pastel_blue_fill),
            "high": ("tolerant", pastel_pink_fill),
        }
    }
    SUCROSE_LIKE_SHEETS = set(SUCROSE_LIKE_MAP.keys())

    # ---- MAPOVÁNÍ pro listy, kde rozhoduje jen přítomnost ALT ----
    # (Saponins je řešené speciální logikou níže -> proto není v tomto mapování)
    ALT_PRESENCE_MAP = {
        "Aphid_Resistance_Rag2": {
            "alt_min": 1,
            "alt": ("Resistant", light_green_fill),
            "ref": ("susceptible", None),
        },
        "Aromatic": {
            "alt_min": 1,
            "alt": ("Aromatic", light_green_fill),
            "ref": ("normal", None),
        },
        "Carotenoid": {
            "alt_min": 1,
            "alt": ("Increased", light_green_fill),
            "ref": ("normal", None),
        },
        "Drought_resistance": {
            "alt_min": 1,
            "alt": ("predicted_as_resistant", light_green_fill),
            "ref": ("normal", None),
        },
        "Flower_color": {
            "alt_min": 1,
            "alt": ("white", None),
            "ref": ("purple", pastel_purple_fill),
        },
        "Glabrous": {
            "alt_min": 1,
            "alt": ("glabrous", light_green_fill),
            "ref": ("with trichomes", None),
        },
        "Hard_seed": {
            "alt_min": 1,
            "alt": ("non-hard seed", light_green_fill),
            "ref": ("hard seed", None),
        },
        "Narrow_leaves_3_seeded_pods": {
            "alt_min": 1,
            "alt": ("narrow leaves", light_green_fill),
            "ref": ("normal", None),
        },
        "Seed_coat_Brown_RedBrown": {
            "alt_min": 1,
            "alt": ("likely red-brown", pastel_pink_fill),
            "ref": ("likely brown", brown_fill),
        },
        "Seed_coat_Green": {
            "alt_min": 1,
            "alt": ("likely green", light_green_fill),
            "ref": ("non green", None),
        },
        "Seed_coat_luster": {
            "alt_min": 1,
            "alt": ("shiny", pastel_purple_fill),
            "ref": ("dull", light_black_fill),
        },
        "Semi-determinate_stem": {
            "alt_min": 1,
            "alt": ("semi-determinate (dependent on Dt1)", light_green_fill),
            "ref": ("non-semi-determinate", None),
        },
        "Stay_Green": {
            "alt_min": 1,
            "alt": ("green", light_green_fill),
            "ref": ("non green", None),
        },
        "Leaf-chewing_insect_resistance": {
            "alt_min": 1,
            "alt": ("resistant", light_green_fill),
            "ref": ("non-resistant", None),
        },
        "Lipoxygenase": {
            "alt_min": 1,
            "alt": ("potentially lower lipoxygenase activity", light_green_fill),
            "ref": ("normal", None),
        },
        "FA_content": {
            "alt_min": 1,
            "alt": ("modulated FA ratios", light_green_fill),
            "ref": ("normal", None),
        },
    }
    ALT_PRESENCE_SHEETS = set(ALT_PRESENCE_MAP.keys())

    wb = load_workbook(vstupni_soubor, data_only=False)

    def norm(v):
        if v is None:
            return ""
        return str(v).strip()

    # Stem_termination pozice
    STEM_DET_POS = {"45125856", "45128025", "45292968"}
    STEM_TALL_POS = {"45177229", "45225588", "45368157", "45184804"}

    # pozice pro Cadmium_accumulation
    CAD_POS_A = "4966222"
    CAD_POS_B = "40474438"

    # pozice pro Fatty_acid
    # FATTY_OLEIC_POS = {"48394045", "50014632", "50370234", "35165204", "35169915"}

    # pozice pro Carbohydrates
    CARB_RS3_POS = {"307744", "311546", "308412"}
    CARB_SS_POS = "47038305"
    CARB_RS2_POS = "15222108"

    # pozice pro Pubsecence_color
    PUB_T_POS = {
        "18926533", "18985125", "19005569", "19020804", "18731376",
        "17812261", "18048014", "19086527", "18737806", "18902586", "18737324"
    }
    PUB_TD_POS = {"45301350", "45301379", "45301305", "45301308"}

    # Pod_shatter klíčové pozice
    POD_POS_A = "29944393"
    POD_POS_B = "1727642"

    # Pod_color pozice
    PODCOLOR_A_POS = {"37819390", "37806119", "37819382", "37819253", "37806091", "37806160"}
    PODCOLOR_B_POS = {
        "538578", "545476", "547447", "922303", "333443", "336617", "337131", "339421",
        "512215", "517157", "520318", "521055", "525858", "528046", "542804"
    }

    # SCN_resistance – MAIN pozice
    SCN_MAIN_POS = {
        "8367180", "8367826", "8319968", "8350676", "8356824",
        "32968127", "32969916",
        "1643660", "1645400", "1624870", "1660502", "1635149",
        "1634135", "1644284"
    }

    # SCN_resistance – MOD pozice
    SCN_MOD_POS = {
        "8367180", "8367826", "8319968", "8350676", "8356824",
        "32968127", "32969916", "1634135", "1644284"
    }

    # Protein_Oil – sledované pozice
    PROTEIN_OIL_POS = {
        "4951338", "5862141", "7463176", "2026371", "5025509",
        "5490549", "5954852", "3875101", "33119281"
    }
    PROTEIN_OIL_TRIGGER_POS = "4951338"  # pokud ALT -> Oil_down_Protein_up

    # Protein_oil_content – pozice
    POIC_POS_A = "3875101"
    POIC_POS_B = "33119281"

    # Seed_coat_Black_Brown – pozice
    SEEDCOAT_BB_POS = {
        "45759100", "45655999", "45662671", "45734775", "45759137",
        "45823324", "45868530", "45905417"
    }

    # Temperature_variations,_phytoch – pozice
    TEMPV_POS_A = "47397774"
    TEMPV_POS_B = "47398384"

    # Saponins – nové pravidlo
    SAPO_MAIN_POS = {
        "43139148", "43139151", "43139150",
        "5338642", "3728854", "8075111", "4361148",
        "46461966", "46141261", "46539089"
    }
    SAPO_POS_SPECIAL = "137342"

    # Big_seed_Protein_Oil – pozice
    BIG_SEED_PO_POS = {"3868776", "3872139", "3868730", "3940118", "31795724"}

    # Internode_length – pozice
    INTERNODE_A_POS = {"37799084", "37809273", "37811772", "37757704"}
    INTERNODE_B_POS = "38414925"

    
    for ws in wb.worksheets:
        max_row = ws.max_row
        max_col = ws.max_column

        # --- 1) najdi řádek 'variant_position' v 1. sloupci ---
        VARIANT_ROW = None
        for rr in range(1, max_row + 1):
            v = ws.cell(row=rr, column=1).value
            if v and norm(v).lower() == "variant_position":
                VARIANT_ROW = rr
                break

        if VARIANT_ROW is None:
            print(f"⚠ List {ws.title} nemá řádek 'variant_position' → přeskočeno.")
            continue

        # --- 2) vytvoř nový sloupec phenotype ---
        new_col = max_col + 1
        new_col_letter = get_column_letter(new_col)

        last_letter = get_column_letter(max_col)
        if last_letter in ws.column_dimensions:
            ws.column_dimensions[new_col_letter].width = ws.column_dimensions[last_letter].width

        hcell = ws.cell(row=HEADER_ROW, column=new_col)
        hcell.value = "phenotype"
        hcell.font = Font(bold=True)
        hcell.border = cell_border

        for rr in range(HEADER_ROW + 1, max_row + 1):
            ws.cell(row=rr, column=new_col).border = cell_border

        # --- NOVĚ: projdi celý 7. řádek a najdi sloupce "Status" ---
        status_cols = []
        for cc in range(1, max_col + 1):
            header_val = norm(ws.cell(row=HEADER_ROW, column=cc).value).lower()
            if header_val == "status":
                status_cols.append(cc)

        # --- 3) zpracování datových řádků ---
        for r in range(FIRST_DATA_ROW, max_row + 1):

        # --- NOVĚ: Status priorita (pokud kdekoliv "unknown" => phenotype unknown)
        # Výjimka: neaplikovat na: 
            forced_unknown = False
            if ws.title not in ("Flowering_time_&_maturity", "Saponins", "SCN_resistance","Pod_shatter", "FA_content"):
                for sc in status_cols:
                    st = norm(ws.cell(row=r, column=sc).value).lower()
                    if st == "unknown":
                        forced_unknown = True
                        break

            if forced_unknown:
                cell = ws.cell(row=r, column=new_col)
                cell.value = "unknown"
                cell.fill = grey_fill
                continue

            ref_count = 0
            alt_count = 0
            blank_count = 0  # používá se smysluplně jen pro sucrose-like

            # Pod_shatter: stav dvou klíčových pozic
            pod_299_state = None  # "ref"/"alt"/None
            pod_172_state = None  # "ref"/"alt"/None

            # Stem_termination
            stem_det_any_alt = False
            stem_tall_any_alt = False
            stem_any_called = False

            # Cadmium_accumulation
            cad_a_state = None
            cad_b_state = None

            # Fatty_acid
            # fat_oleic_any_called = False
            # fat_oleic_any_alt = False
            # fat_oleic_any_ref = False

            # Carbohydrates
            carb_rs3_any_alt = False
            carb_rs3_all_ref = True
            carb_ss_state = None
            carb_rs2_state = None

            # Pubsecence_color (upravené pro ignorování prázdných buněk)
            pub_t_all_ref = True
            pub_t_any_alt = False
            pub_t_any_called = False

            pub_td_all_ref = True
            pub_td_any_alt = False
            pub_td_any_called = False

            # Pod_color
            podc_a_any_alt = False
            podc_a_all_ref = True
            podc_a_any_called = False

            podc_b_any_alt = False
            podc_b_all_ref = True
            podc_b_any_called = False

            # SCN_resistance – MAIN
            scn_main_any_called = False
            scn_main_any_alt = False
            scn_main_any_ref = False

            # SCN_resistance – MOD
            scn_mod_any_called = False
            scn_mod_any_alt = False
            scn_mod_all_alt = True

            # Protein_Oil (upraveno)
            po_any_called = False
            po_all_ref = True
            po_trigger_495_alt = False

            # Protein_oil_content
            poic_a_state = None
            poic_b_state = None

            # Seed_coat_Black_Brown
            scbb_any_called = False
            scbb_any_alt = False
            scbb_any_ref = False

            # Temperature_variations,_phytoch
            tempv_a_state = None
            tempv_b_state = None

            # Saponins
            sapo_main_all_called = True
            sapo_main_all_ref = True
            sapo_main_any_alt = False
            sapo_special_state = None

            # Big_seed_Protein_Oil
            bs_po_any_called = False
            bs_po_any_ref = False
            bs_po_all_alt = True

            # Internode_length
            internode_a_any_alt = False
            internode_a_called = set()   # bude obsahovat 4 pozice, pokud jsou všechny callnuté REF/ALT
            internode_a_bad = False      # pokud narazíme na het/n/alt2/unknown hodnotu v A, zneplatníme

            internode_b_state = None     # "ref"/"alt"/None
            internode_b_bad = False      # het/n/alt2/unknown v B => unknown


            for c in range(1, max_col + 1):
                pos_val = norm(ws.cell(row=VARIANT_ROW, column=c).value)
                if not pos_val:
                    continue

                if pos_val in EXCEPTIONS:
                    ref_name, alt_name = EXCEPTIONS[pos_val]
                elif pos_val in RULES:
                    ref_name, alt_name = RULES[pos_val]
                else:
                    continue

                ref_name = "" if ref_name is None else norm(ref_name)
                alt_name = "" if alt_name is None else norm(alt_name)

                value = norm(ws.cell(row=r, column=c).value)

                # --- PRÁZDNÉ BUŇKY: ignorovat úplně (kromě sucrose-like, kde se počítají) ---
                if value == "":
                    if ws.title in SUCROSE_LIKE_SHEETS:
                        blank_count += 1
                    continue

                # Het / N / alt2
                # (ponecháno jak jsi měla: u sucrose-like se berou do blank_count; jinde ovlivňují "called" logiku tam, kde je potřeba)
                if value.lower() in ("het", "n", "alt2"):
                    if ws.title in SUCROSE_LIKE_SHEETS:
                        blank_count += 1

                    # Pubsecence_color: necalled (ale prázdno ignorujeme; het/n/alt2 je explicitně necalled)
                    if ws.title == "Pubsecence_color":
                        if pos_val in PUB_T_POS:
                            # nic do all_ref/any_alt, jen nepočítat jako called
                            pass
                        elif pos_val in PUB_TD_POS:
                            pass

                    # SCN_resistance: het/n/alt2 = necalled, ruší "all alt" pro MOD
                    if ws.title == "SCN_resistance" and pos_val in SCN_MOD_POS:
                        scn_mod_all_alt = False

                    # Protein_Oil: het/n/alt2 ruší "vše REF"
                    if ws.title == "Protein_Oil" and pos_val in PROTEIN_OIL_POS:
                        po_all_ref = False

                    # Saponins: het/n/alt2 = necalled
                    if ws.title == "Saponins":
                        if pos_val in SAPO_MAIN_POS:
                            sapo_main_all_called = False
                            sapo_main_all_ref = False
                        elif pos_val == SAPO_POS_SPECIAL:
                            sapo_special_state = None

                    # Big_seed_Protein_Oil
                    if ws.title == "Big_seed_Protein_Oil" and pos_val in BIG_SEED_PO_POS:
                        bs_po_all_alt = False
                        bs_po_any_called = True

                    # Internode_length: het/n/alt2 = necallnuté => unknown
                    if ws.title == "Internode_length":
                        if pos_val in INTERNODE_A_POS:
                            internode_a_bad = True
                        elif pos_val == INTERNODE_B_POS:
                            internode_b_bad = True
                            internode_b_state = None


                    continue

                # ALT
                if alt_name and value == alt_name:
                    alt_count += 1

                    if ws.title == "Pod_shatter":
                        if pos_val == POD_POS_A:
                            pod_299_state = "alt"
                        elif pos_val == POD_POS_B:
                            pod_172_state = "alt"

                    if ws.title == "Stem_termination":
                        if pos_val in STEM_DET_POS:
                            stem_det_any_alt = True
                            stem_any_called = True
                        elif pos_val in STEM_TALL_POS:
                            stem_tall_any_alt = True
                            stem_any_called = True

                    if ws.title == "Cadmium_accumulation":
                        if pos_val == CAD_POS_A:
                            cad_a_state = "alt"
                        elif pos_val == CAD_POS_B:
                            cad_b_state = "alt"

                    # if ws.title == "FA_content" and pos_val in FATTY_OLEIC_POS:
                    #     fat_oleic_any_called = True
                    #     fat_oleic_any_alt = True

                    if ws.title == "Carbohydrates":
                        if pos_val in CARB_RS3_POS:
                            carb_rs3_any_alt = True
                            carb_rs3_all_ref = False
                        elif pos_val == CARB_SS_POS:
                            carb_ss_state = "alt"
                        elif pos_val == CARB_RS2_POS:
                            carb_rs2_state = "alt"

                    if ws.title == "Pubsecence_color":
                        if pos_val in PUB_T_POS:
                            pub_t_any_called = True
                            pub_t_any_alt = True
                            pub_t_all_ref = False
                        elif pos_val in PUB_TD_POS:
                            pub_td_any_called = True
                            pub_td_any_alt = True
                            pub_td_all_ref = False

                    if ws.title == "Pod_color":
                        if pos_val in PODCOLOR_A_POS:
                            podc_a_any_called = True
                            podc_a_any_alt = True
                            podc_a_all_ref = False
                        elif pos_val in PODCOLOR_B_POS:
                            podc_b_any_called = True
                            podc_b_any_alt = True
                            podc_b_all_ref = False

                    if ws.title == "SCN_resistance":
                        if pos_val in SCN_MAIN_POS:
                            scn_main_any_called = True
                            scn_main_any_alt = True
                        if pos_val in SCN_MOD_POS:
                            scn_mod_any_called = True
                            scn_mod_any_alt = True

                    if ws.title == "Protein_Oil" and pos_val in PROTEIN_OIL_POS:
                        po_any_called = True
                        po_all_ref = False
                        if pos_val == PROTEIN_OIL_TRIGGER_POS:
                            po_trigger_495_alt = True

                    if ws.title == "Protein_oil_content":
                        if pos_val == POIC_POS_A:
                            poic_a_state = "alt"
                        elif pos_val == POIC_POS_B:
                            poic_b_state = "alt"

                    if ws.title == "Seed_coat_Black_Brown" and pos_val in SEEDCOAT_BB_POS:
                        scbb_any_called = True
                        scbb_any_alt = True

                    if ws.title == "Temperature_variations,_phytoch":
                        if pos_val == TEMPV_POS_A:
                            tempv_a_state = "alt"
                        elif pos_val == TEMPV_POS_B:
                            tempv_b_state = "alt"

                    if ws.title == "Saponins":
                        if pos_val in SAPO_MAIN_POS:
                            sapo_main_any_alt = True
                            sapo_main_all_ref = False
                        elif pos_val == SAPO_POS_SPECIAL:
                            sapo_special_state = "alt"

                    if ws.title == "Big_seed_Protein_Oil" and pos_val in BIG_SEED_PO_POS:
                        bs_po_any_called = True
                        # ALT podporuje "all_alt" (nemění)

                    if ws.title == "Internode_length":
                        if pos_val in INTERNODE_A_POS:
                            internode_a_any_alt = True
                            internode_a_called.add(pos_val)
                        elif pos_val == INTERNODE_B_POS:
                            internode_b_state = "alt"


                    continue

                # REF
                if ref_name and value == ref_name:
                    ref_count += 1

                    if ws.title == "Pod_shatter":
                        if pos_val == POD_POS_A:
                            pod_299_state = "ref"
                        elif pos_val == POD_POS_B:
                            pod_172_state = "ref"

                    if ws.title == "Stem_termination":
                        if pos_val in STEM_DET_POS or pos_val in STEM_TALL_POS:
                            stem_any_called = True

                    if ws.title == "Cadmium_accumulation":
                        if pos_val == CAD_POS_A:
                            cad_a_state = "ref"
                        elif pos_val == CAD_POS_B:
                            cad_b_state = "ref"

                    # if ws.title == "FA_content" and pos_val in FATTY_OLEIC_POS:
                    #     fat_oleic_any_called = True
                    #     fat_oleic_any_ref = True
                    if ws.title == "Carbohydrates":
                        if pos_val == CARB_SS_POS:
                            carb_ss_state = "ref"
                        elif pos_val == CARB_RS2_POS:
                            carb_rs2_state = "ref"

                    if ws.title == "Pubsecence_color":
                        if pos_val in PUB_T_POS:
                            pub_t_any_called = True
                            # all_ref zůstává True
                        elif pos_val in PUB_TD_POS:
                            pub_td_any_called = True

                    if ws.title == "Pod_color":
                        if pos_val in PODCOLOR_A_POS:
                            podc_a_any_called = True
                        elif pos_val in PODCOLOR_B_POS:
                            podc_b_any_called = True

                    if ws.title == "SCN_resistance":
                        if pos_val in SCN_MAIN_POS:
                            scn_main_any_called = True
                            scn_main_any_ref = True
                        if pos_val in SCN_MOD_POS:
                            scn_mod_any_called = True
                            scn_mod_all_alt = False

                    if ws.title == "Protein_Oil" and pos_val in PROTEIN_OIL_POS:
                        po_any_called = True
                        # po_all_ref zůstává True jen pokud nikde nebyla alt/het/unknown v PROTEIN_OIL_POS

                    if ws.title == "Protein_oil_content":
                        if pos_val == POIC_POS_A:
                            poic_a_state = "ref"
                        elif pos_val == POIC_POS_B:
                            poic_b_state = "ref"

                    if ws.title == "Seed_coat_Black_Brown" and pos_val in SEEDCOAT_BB_POS:
                        scbb_any_called = True
                        scbb_any_ref = True

                    if ws.title == "Temperature_variations,_phytoch":
                        if pos_val == TEMPV_POS_A:
                            tempv_a_state = "ref"
                        elif pos_val == TEMPV_POS_B:
                            tempv_b_state = "ref"

                    if ws.title == "Saponins":
                        if pos_val in SAPO_MAIN_POS:
                            # ref -> nic
                            pass
                        elif pos_val == SAPO_POS_SPECIAL:
                            sapo_special_state = "ref"

                    if ws.title == "Big_seed_Protein_Oil" and pos_val in BIG_SEED_PO_POS:
                        bs_po_any_called = True
                        bs_po_any_ref = True
                        bs_po_all_alt = False

                    if ws.title == "Internode_length":
                        if pos_val in INTERNODE_A_POS:
                            internode_a_called.add(pos_val)
                        elif pos_val == INTERNODE_B_POS:
                            internode_b_state = "ref"

                    continue

                # --- hodnota není ani REF ani ALT => unknown v této pozici ---
                if ws.title == "Carbohydrates" and pos_val in CARB_RS3_POS:
                    carb_rs3_all_ref = False

                if ws.title == "Pod_color":
                    if pos_val in PODCOLOR_A_POS:
                        podc_a_all_ref = False
                    elif pos_val in PODCOLOR_B_POS:
                        podc_b_all_ref = False

                if ws.title == "SCN_resistance" and pos_val in SCN_MOD_POS:
                    scn_mod_all_alt = False

                if ws.title == "Protein_Oil" and pos_val in PROTEIN_OIL_POS:
                    po_all_ref = False
                    po_any_called = True

                if ws.title == "Seed_coat_Black_Brown" and pos_val in SEEDCOAT_BB_POS:
                    scbb_any_called = True

                if ws.title == "Saponins":
                    if pos_val in SAPO_MAIN_POS:
                        sapo_main_all_called = False
                        sapo_main_all_ref = False
                    elif pos_val == SAPO_POS_SPECIAL:
                        sapo_special_state = None

                if ws.title == "Internode_length":
                    if pos_val in INTERNODE_A_POS:
                        internode_a_bad = True
                    elif pos_val == INTERNODE_B_POS:
                        internode_b_bad = True
                        internode_b_state = None

            # --- 4) zapiš phenotype výsledek ---
            cell = ws.cell(row=r, column=new_col)
            cell.fill = PatternFill(fill_type=None)

            # Pod_shatter (nová pravidla)
            if ws.title == "Pod_shatter":
                # 1) Všechny geny ALT => non-shatering
                if alt_count > 0 and ref_count == 0:
                    cell.value = "non-shatering"

                # 2) Všechny geny REF => shatering
                elif ref_count > 0 and alt_count == 0:
                    cell.value = "shatering"

                # 3) Jinak rozhodují klíčové pozice 29944393 a 1727642
                else:
                    # a) 299 = ref, 172 = ref/alt/nevíme => predicted as shatering
                    if pod_299_state == "ref":
                        cell.value = "predicted as shatering"
                        cell.fill = light_green_fill

                    # b) 299 = alt
                    elif pod_299_state == "alt":
                        # 299 alt + 172 alt => predicted as non-shatering
                        if pod_172_state == "alt":
                            cell.value = "predicted as non-shatering"
                        # 299 alt + 172 ref => predicted as shatering
                        elif pod_172_state == "ref":
                            cell.value = "predicted as shatering"
                            cell.fill = light_green_fill
                        # 299 alt + 172 nevíme => unknown
                        else:
                            cell.value = "unknown"
                            cell.fill = grey_fill

                    # c) 299 nevíme (a nejsme v all-alt / all-ref) => unknown
                    else:
                        cell.value = "unknown"
                        cell.fill = grey_fill

            # Pod_color
            elif ws.title == "Pod_color":
                if not (podc_a_any_called and podc_b_any_called):
                    cell.value = "unknown"
                    cell.fill = grey_fill
                else:
                    if podc_a_any_alt and podc_b_any_alt:
                        cell.value = "Tan"
                        cell.fill = tan_fill
                    elif podc_a_any_alt and podc_b_all_ref:
                        cell.value = "Brown"
                        cell.fill = brown_fill
                    elif podc_a_all_ref and podc_b_any_alt:
                        cell.value = "black"
                        cell.fill = light_black_fill
                    elif podc_a_all_ref and podc_b_all_ref:
                        cell.value = "black"
                        cell.fill = light_black_fill
                    else:
                        cell.value = "unknown"
                        cell.fill = grey_fill

            # Cadmium_accumulation
            elif ws.title == "Cadmium_accumulation":
                if cad_a_state in ("ref", "alt") and cad_b_state in ("ref", "alt"):
                    if (cad_a_state == "ref" and cad_b_state == "ref") or (cad_a_state == "alt" and cad_b_state == "alt"):
                        cell.value = "high/low"
                        cell.fill = pastel_purple_fill
                    elif cad_a_state == "ref" and cad_b_state == "alt":
                        cell.value = "high"
                        cell.fill = pastel_pink_fill
                    elif cad_a_state == "alt" and cad_b_state == "ref":
                        cell.value = "low"
                        cell.fill = pastel_blue_fill
                    else:
                        cell.value = "unknown"
                        cell.fill = grey_fill
                else:
                    cell.value = "unknown"
                    cell.fill = grey_fill

            # Fatty_acid – nové pravidlo (oleic acid)
            # elif ws.title == "FA_content":
            #     if not fat_oleic_any_called:
            #         cell.value = "unknown"
            #         cell.fill = grey_fill
            #     elif fat_oleic_any_alt:
            #         cell.value = "slightly increased oleic acid"
            #         cell.fill = pastel_blue_fill
            #     elif fat_oleic_any_ref:
            #         cell.value = "high oleic acid"
            #         cell.fill = pastel_pink_fill
            #     else:
            #         cell.value = "unknown"
            #         cell.fill = grey_fill

            # SCN_resistance
            elif ws.title == "SCN_resistance":
                if scn_main_any_called and scn_main_any_ref and (not scn_main_any_alt):
                    cell.value = "susceptible"
                elif scn_main_any_called and scn_main_any_alt and (not scn_main_any_ref):
                    cell.value = "resistant"
                    cell.fill = light_green_fill
                elif scn_mod_any_called and scn_mod_all_alt:
                    cell.value = "moderately resistant"
                    cell.fill = pastel_blue_fill
                elif scn_mod_any_alt:
                    cell.value = "moderately susceptible"
                    cell.fill = pastel_pink_fill
                else:
                    cell.value = "unknown"
                    cell.fill = grey_fill

            # Carbohydrates
            elif ws.title == "Carbohydrates":
                if carb_rs3_all_ref and (carb_ss_state == "ref") and (carb_rs2_state == "ref"):
                    cell.value = "normal"
                elif carb_rs3_any_alt and (carb_ss_state == "ref") and (carb_rs2_state in ("ref", "alt")):
                    cell.value = "low raffinose"
                    cell.fill = pastel_blue_fill
                elif carb_rs3_all_ref and (carb_ss_state == "alt") and (carb_rs2_state in ("ref", "alt")):
                    cell.value = "low stachyose"
                    cell.fill = pastel_pink_fill
                elif carb_rs3_any_alt and (carb_ss_state == "alt") and (carb_rs2_state in ("ref", "alt")):
                    cell.value = "lower contain."
                    cell.fill = light_green_fill
                else:
                    cell.value = "unknown"
                    cell.fill = grey_fill

            # Pubsecence_color (upravené)
            elif ws.title == "Pubsecence_color":
                # pokud nemáme nic callnuté v obou skupinách, nedává smysl určovat barvu
                if not (pub_t_any_called and pub_td_any_called):
                    cell.value = "unknown"
                    cell.fill = grey_fill
                else:
                    if pub_t_all_ref and (not pub_t_any_alt) and pub_td_all_ref and (not pub_td_any_alt):
                        cell.value = "tawny"
                        cell.fill = tawny_fill
                    elif pub_t_all_ref and (not pub_t_any_alt) and pub_td_any_alt:
                        cell.value = "light_tawny"
                        cell.fill = light_tawny_fill
                    elif pub_t_any_alt:
                        cell.value = "grey"
                    else:
                        cell.value = "unknown"
                        cell.fill = grey_fill

            # Stem_termination
            elif ws.title == "Stem_termination":
                if stem_det_any_alt and stem_tall_any_alt:
                    cell.value = "determinant/tall determinant"
                    cell.fill = pastel_pink_fill
                elif stem_tall_any_alt:
                    cell.value = "tall determinant"
                    cell.fill = light_green_fill
                elif stem_det_any_alt:
                    cell.value = "determinant"
                    cell.fill = light_green_fill
                else:
                    if stem_any_called:
                        cell.value = "indeterminant"
                    else:
                        cell.value = "unknown"
                        cell.fill = grey_fill

            # Flowering_time_&_maturity
            elif ws.title == "Flowering_time_&_maturity":
                total_called = ref_count + alt_count
                if total_called == 0:
                    cell.value = "unknown"
                    cell.fill = grey_fill
                else:
                    percent_alt = (alt_count * 100.0) / total_called
                    if percent_alt >= 81:
                        cell.value = "late"
                        cell.fill = ultra_early_fill
                    elif percent_alt >= 61:
                        cell.value = "mid-early"
                        cell.fill = very_early_fill
                    elif percent_alt >= 41:
                        cell.value = "early"
                        cell.fill = early_fill
                    elif percent_alt >= 21:
                        cell.value = "very early"
                        cell.fill = mid_early_fill
                    else:
                        cell.value = "extra-early"
                        cell.fill = late_fill

            # Sucrose-like
            elif ws.title in SUCROSE_LIKE_SHEETS:
                total = ref_count + alt_count + blank_count
                if total > 0:
                    percent_alt = round((alt_count * 100.0) / total, 2)
                    cfg = SUCROSE_LIKE_MAP[ws.title]
                    mid_max = float(cfg.get("mid_max", 50.0))
                    mid_inclusive = bool(cfg.get("mid_inclusive", True))

                    if percent_alt == 0.0:
                        label, fill = cfg["zero"]
                    else:
                        in_mid = (percent_alt <= mid_max) if mid_inclusive else (percent_alt < mid_max)
                        label, fill = cfg["mid"] if in_mid else cfg["high"]

                    cell.value = label
                    if fill is not None:
                        cell.fill = fill
                else:
                    cell.value = None

                if cell.value == "unknown":
                    cell.fill = grey_fill

            # Saponins
            elif ws.title == "Saponins":
                if sapo_special_state == "alt" and sapo_main_all_called and sapo_main_all_ref:
                    cell.value = "modulated ratio of saponins"
                    cell.fill = light_green_fill
                elif sapo_special_state == "alt" and sapo_main_any_alt:
                    cell.value = "lower"
                    cell.fill = pastel_pink_fill
                elif sapo_special_state == "ref" and sapo_main_any_alt:
                    cell.value = "lower"
                    cell.fill = light_green_fill
                elif sapo_special_state == "ref" and sapo_main_all_called and sapo_main_all_ref:
                    cell.value = "normal"
                else:
                    cell.value = "unknown"
                    cell.fill = grey_fill

            # Protein_Oil
            elif ws.title == "Protein_Oil":
                if po_trigger_495_alt:
                    cell.value = "Oil_down_Protein_up"
                    cell.fill = pastel_blue_fill
                elif po_any_called and po_all_ref:
                    cell.value = "normal"
                else:
                    cell.value = "unknown"
                    cell.fill = grey_fill

            # Protein_oil_content
            elif ws.title == "Protein_oil_content":
                if poic_a_state in ("ref", "alt") and poic_b_state in ("ref", "alt"):
                    if poic_a_state == "ref":
                        cell.value = "Oil_up_Protein_down"
                        cell.fill = pastel_pink_fill
                    elif poic_a_state == "alt":
                        cell.value = "normal"
                    else:
                        cell.value = "unknown"
                        cell.fill = grey_fill
                else:
                    cell.value = "unknown"
                    cell.fill = grey_fill

            # Seed_coat_Black_Brown
            elif ws.title == "Seed_coat_Black_Brown":
                if scbb_any_alt:
                    cell.value = "non-black"
                elif scbb_any_called and scbb_any_ref:
                    cell.value = "black"
                    cell.fill = light_black_fill
                else:
                    cell.value = "unknown"
                    cell.fill = grey_fill

            # Temperature_variations,_phytoch
            elif ws.title == "Temperature_variations,_phytoch":
                if tempv_a_state in ("ref", "alt") and tempv_b_state in ("ref", "alt"):
                    if tempv_a_state == "ref" and tempv_b_state == "ref":
                        cell.value = "normal"
                    elif tempv_a_state == "alt" and tempv_b_state == "alt":
                        cell.value = "chilling tolerance"
                        cell.fill = light_green_fill
                    elif tempv_a_state == "alt" and tempv_b_state == "ref":
                        cell.value = "slightly delays flowering"
                        cell.fill = pastel_blue_fill
                    elif tempv_a_state == "ref" and tempv_b_state == "alt":
                        cell.value = "delays flowering"
                        cell.fill = pastel_pink_fill
                    else:
                        cell.value = "unknown"
                        cell.fill = grey_fill
                else:
                    cell.value = "unknown"
                    cell.fill = grey_fill

            # Big_seed_Protein_Oil
            elif ws.title == "Big_seed_Protein_Oil":
                if bs_po_any_ref:
                    cell.value = "Oil_up_Protein_down"
                    cell.fill = pastel_pink_fill
                elif bs_po_any_called and bs_po_all_alt:
                    cell.value = "Oil_down_Protein_up"
                    cell.fill = pastel_blue_fill
                else:
                    cell.value = "unknown"
                    cell.fill = grey_fill
            
             # Internode_length (nová logika podle kombinací A/B)
            elif ws.title == "Internode_length":
                # A je platné jen pokud jsou všechny 4 pozice callnuté jako REF/ALT a nebyl problém
                if (not internode_a_bad) and (len(internode_a_called) == len(INTERNODE_A_POS)):
                    internode_a_state = "alt" if internode_a_any_alt else "ref"
                else:
                    internode_a_state = None

                # B je platné jen pokud je ref/alt a nebyl problém
                if internode_b_bad:
                    internode_b_state = None

                if internode_a_state in ("ref", "alt") and internode_b_state in ("ref", "alt"):
                    if internode_a_state == "alt" and internode_b_state == "ref":
                        cell.value = "likely non-elongated plants"
                    elif internode_a_state == "alt" and internode_b_state == "alt":
                        cell.value = "likely shorter compact plants"
                        cell.fill = light_green_fill
                    elif internode_a_state == "ref" and internode_b_state == "alt":
                        cell.value = "likely slightly elongated plants"
                        cell.fill = pastel_blue_fill
                    elif internode_a_state == "ref" and internode_b_state == "ref":
                        cell.value = "likely elongated plants"
                        cell.fill = pastel_pink_fill
                    else:
                        cell.value = "unknown"
                        cell.fill = grey_fill
                else:
                    cell.value = "unknown"
                    cell.fill = grey_fill

            # ALT presence
            elif ws.title in ALT_PRESENCE_SHEETS:
                cfg = ALT_PRESENCE_MAP[ws.title]
                if alt_count >= int(cfg.get("alt_min", 1)):
                    label, fill = cfg["alt"]
                else:
                    label, fill = cfg["ref"]

                cell.value = label
                if fill is not None:
                    cell.fill = fill

            # Obecná logika
            else:
                if alt_count == 0 and ref_count > 0:
                    cell.value = "100% functional"
                elif ref_count == 0 and alt_count > 0:
                    cell.value = "100% non-functional"
                elif alt_count > ref_count:
                    cell.value = "likely non-functional"
                elif ref_count > alt_count:
                    cell.value = "likely functional"
                else:
                    cell.value = "unknown"

    # --- uložení ---
    out_folder = Path("phenotype_evaluation")
    out_folder.mkdir(parents=True, exist_ok=True)
    out_path = out_folder / "phenotype.xlsx"
    wb.save(out_path)

    print(f"Hotovo! Vytvořen soubor: {out_path}")
