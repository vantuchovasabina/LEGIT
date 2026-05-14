from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from copy import copy

def reorder_marker_columns(vstupni_soubor: str):
    "„Funkce fc_marker_column načte vstupní Excel soubor a na každém listu přesune sloupce označené v řádku 7 hodnotou Marker na konec jejich příslušné sekce. "
    "Sekce jsou určeny podle sloučených buněk ve 2. řádku (pokud nejsou, bere se celý list jako jedna sekce). Přeskupení provede tak, "
    "že vytvoří dočasný list se stejným obsahem, zachová hodnoty, formátování, šířky sloupců, výšky řádků, sloučené buňky i freeze panes, a poté původní list nahradí novým."
    OUT_FILE = "All_Traits_matrix_reordered.xlsx"
    out_folder = Path("reordered_matrices")
    out_folder.mkdir(parents=True, exist_ok=True)
    out_path = out_folder / OUT_FILE

    wb = load_workbook(vstupni_soubor, data_only=False)
    HEADER_ROW = 7

    def norm(val):
        if val is None:
            return ""
        return str(val).strip()

    def is_marker(val):
        return norm(val).lower() == "marker"

    for ws in list(wb.worksheets):
        max_row = ws.max_row
        max_col = ws.max_column

        if max_row < HEADER_ROW:
            print(f"List '{ws.title}' přeskočen: nemá {HEADER_ROW} řádků.")
            continue

        header_vals = [ws.cell(row=HEADER_ROW, column=c).value for c in range(1, max_col + 1)]

        # --- Sekce podle sloučených buněk na 2. řádku (jak máš v kódu) ---
        sections = []
        merged_ranges_r2 = [
            mr for mr in ws.merged_cells.ranges if mr.min_row == 2 and mr.max_row >= 2
        ]

        if not merged_ranges_r2:
            if max_col > 0:
                sections.append((1, max_col))
            print(f"List '{ws.title}': Sekce definována jako 1 až {max_col} (bez sloučených buněk na řádku 2).")
        else:
            for mr in merged_ranges_r2:
                sections.append((mr.min_col, mr.max_col))
            sections.sort(key=lambda x: x[0])
            final_sections = []
            current_col = 1
            for start_col, end_col in sections:
                if start_col > current_col:
                    final_sections.append((current_col, start_col - 1))
                final_sections.append((start_col, end_col))
                current_col = end_col + 1
            if current_col <= max_col:
                final_sections.append((current_col, max_col))
            sections = final_sections

        print(f"List '{ws.title}': Nalezené sekce: {sections}")

        new_order = list(range(1, max_col + 1))
        change_made = False

        for start, end in sections:
            idxs = list(range(start, end + 1))
            markers = [i for i in idxs if is_marker(header_vals[i-1])]
            others  = [i for i in idxs if not is_marker(header_vals[i-1])]

            if markers:
                change_made = True
                perm = others + markers
                new_order[start-1:end] = perm

        if not change_made:
            print(f"List '{ws.title}' beze změny.")
            continue

        print(f"Zpracovávám list: '{ws.title}' (přeskupení 'Marker' sloupců)...")

        col_map = {old: new_pos for new_pos, old in enumerate(new_order, start=1)}

        # >>>>>>> FIX: vytvoř nový list rovnou na správné pozici a bez _sheets.insert <<<<<<<
        insert_idx = wb.worksheets.index(ws)
        tmp_name = f"{ws.title}__tmp__"
        # pojistka proti kolizi názvů
        while any(sh.title == tmp_name for sh in wb.worksheets):
            tmp_name += "_x"
        ws_new = wb.create_sheet(title=tmp_name, index=insert_idx)

        # šířky sloupců a výšky řádků
        for old_col in range(1, max_col + 1):
            new_col = col_map.get(old_col, old_col)
            col_letter_old = get_column_letter(old_col)
            col_letter_new = get_column_letter(new_col)
            if col_letter_old in ws.column_dimensions:
                ws_new.column_dimensions[col_letter_new].width = ws.column_dimensions[col_letter_old].width

        for r, rd in ws.row_dimensions.items():
            if rd.height is not None:
                ws_new.row_dimensions[r].height = rd.height

        # kopie buněk (hodnoty + styly)
        for old_col in range(1, max_col + 1):
            new_col = col_map.get(old_col, old_col)
            for r in range(1, max_row + 1):
                src = ws.cell(row=r, column=old_col)
                dst = ws_new.cell(row=r, column=new_col)
                dst.value = src.value
                dst.number_format = src.number_format
                if src.has_style:
                    dst.font = copy(src.font)
                    dst.fill = copy(src.fill)
                    dst.border = copy(src.border)
                    dst.alignment = copy(src.alignment)
                    dst.protection = copy(src.protection)

        # freeze panes
        ws_new.freeze_panes = ws.freeze_panes

        # sloučené buňky
        for mr in ws.merged_cells.ranges:
            min_row, min_col, max_row_m, max_col_m = mr.min_row, mr.min_col, mr.max_row, mr.max_col
            original_cols = list(range(min_col, max_col_m + 1))
            new_cols_in_range = [col_map.get(c, c) for c in original_cols]
            new_min_col = min(new_cols_in_range)
            new_max_col = max(new_cols_in_range)
            ws_new.merge_cells(start_row=min_row, start_column=new_min_col,
                               end_row=max_row_m, end_column=new_max_col)

        # smaž původní list a přejmenuj nový na původní název (pořadí zůstane stejné)
        wb.remove(ws)
        ws_new.title = ws.title

    try:
        wb.save(out_path)
    except Exception as e:
        print(f"❌ Chyba při ukládání souboru '{out_path}'. Zavři prosím soubor, pokud je otevřený. Detaily: {e}")
        return

    print(f"Hotovo! Vytvořen soubor: {out_path}")
